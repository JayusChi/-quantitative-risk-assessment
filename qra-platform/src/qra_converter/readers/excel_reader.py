"""Deterministic XLS/XLSX parsing with formulas, merges and hidden metadata."""

from __future__ import annotations

import struct
from datetime import date, datetime
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from ..contracts import IssueSeverity
from ..parsing.compatibility import legacy_read
from ..parsing.contracts import (
    BoundingBox,
    CoordinateSpace,
    ParsedCell,
    ParsedDocument,
    ParsedPage,
    ParsedTable,
    ParseIssue,
    ReaderOutput,
    source_fragment_sha256,
)
from ..parsing.registry import ParseContext


def _display(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, date | datetime):
        return value.isoformat()
    return str(value)


def _value_type(value: object) -> str:
    if value is None:
        return "EMPTY"
    if isinstance(value, bool):
        return "BOOLEAN"
    if isinstance(value, datetime):
        return "DATETIME"
    if isinstance(value, date):
        return "DATE"
    if isinstance(value, int | float):
        return "NUMBER"
    return "STRING"


def _xls_formula_texts(
    workbook: object, sheet_index: int
) -> tuple[dict[tuple[int, int], str], int]:
    """Read BIFF formula tokens without executing them or changing xlrd sheet state."""
    try:
        from xlrd.biffh import XL_EOF, XL_FORMULA_OPCODES
        from xlrd.formula import FMLA_TYPE_CELL, decompile_formula
    except ImportError:  # pragma: no cover
        return {}, 0
    if int(workbook.biff_version) < 30:
        return {}, 0
    position = int(workbook._sh_abs_posn[sheet_index])
    memory = workbook.mem
    formulas: dict[tuple[int, int], str] = {}
    unavailable = 0
    while position + 4 <= len(memory):
        opcode, length = struct.unpack("<HH", memory[position : position + 4])
        position += 4
        data = bytes(memory[position : position + length])
        position += length
        if opcode == XL_EOF:
            break
        if opcode not in XL_FORMULA_OPCODES or len(data) < 22:
            continue
        row_index, column_index = struct.unpack("<HH", data[:4])
        formula_length = struct.unpack("<H", data[20:22])[0]
        if formula_length <= 0 or len(data) < 22 + formula_length:
            unavailable += 1
            continue
        try:
            text = decompile_formula(
                workbook,
                data[22 : 22 + formula_length],
                formula_length,
                FMLA_TYPE_CELL,
                browx=row_index,
                bcolx=column_index,
                r1c1=0,
            )
        except Exception:
            text = None
        if text:
            formulas[(row_index, column_index)] = text if text.startswith("=") else f"={text}"
        else:
            unavailable += 1
    return formulas, unavailable


class XlsxReader:
    reader_id = "xlsx/openpyxl"
    parser_version = "2.0.0"
    media_types = frozenset({"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})

    def parse(self, path: Path, context: ParseContext) -> ReaderOutput:
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.datetime import to_excel
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("读取XLSX需要安装openpyxl>=3.1") from exc
        try:
            formula_book = load_workbook(path, read_only=False, data_only=False, keep_links=False)
            value_book = load_workbook(path, read_only=False, data_only=True, keep_links=False)
        except (BadZipFile, KeyError, OSError, ValueError) as exc:
            raise ValueError(f"XLSX结构无效：{path.name}") from exc
        issues: list[ParseIssue] = []
        tables: list[ParsedTable] = []
        pages: list[ParsedPage] = []
        metadata: dict[str, object] = {}
        try:
            try:
                with ZipFile(path) as package:
                    names = set(package.namelist())
            except BadZipFile:
                names = set()
            embedded = sorted(name for name in names if name.startswith("xl/embeddings/"))
            macros = sorted(name for name in names if name.endswith("vbaProject.bin"))
            external = sorted(name for name in names if name.startswith("xl/externalLinks/"))
            if embedded or macros:
                issues.append(
                    ParseIssue(
                        "PARSE.UNSUPPORTED_EMBEDDED_OBJECT",
                        "工作簿包含未执行的宏或嵌入对象",
                        IssueSeverity.WARNING,
                        location="xl/package",
                    )
                )
            if external:
                issues.append(
                    ParseIssue(
                        "PARSE.UNSUPPORTED_EMBEDDED_OBJECT",
                        "工作簿包含未加载的外部链接",
                        IssueSeverity.WARNING,
                        location="xl/externalLinks",
                    )
                )
            metadata.update(
                {
                    "has_macros": bool(macros),
                    "embedded_object_count": len(embedded),
                    "external_link_count": len(external),
                }
            )
            for page_number, worksheet in enumerate(formula_book.worksheets, start=1):
                if context.cancel_check:
                    context.cancel_check()
                if context.page_progress:
                    context.page_progress(
                        page_number,
                        len(formula_book.worksheets),
                        f"正在解析工作表{worksheet.title}",
                    )
                value_sheet = value_book[worksheet.title]
                max_row = worksheet.max_row if worksheet.max_row and worksheet.max_column else 0
                max_column = worksheet.max_column if max_row else 0
                if max_row * max_column > 2_000_000:
                    raise ValueError(f"工作表{worksheet.title}使用区域超过200万单元格上限")
                merged_by_master: dict[tuple[int, int], tuple[int, int]] = {}
                merged_members: set[tuple[int, int]] = set()
                for merged in worksheet.merged_cells.ranges:
                    master = (merged.min_row, merged.min_col)
                    merged_by_master[master] = (
                        merged.max_row - merged.min_row + 1,
                        merged.max_col - merged.min_col + 1,
                    )
                    for row in range(merged.min_row, merged.max_row + 1):
                        for column in range(merged.min_col, merged.max_col + 1):
                            if (row, column) != master:
                                merged_members.add((row, column))
                if merged_by_master:
                    issues.append(
                        ParseIssue(
                            "PARSE.MERGED_CELL_EXPANDED",
                            f"工作表{worksheet.title}的合并单元格仅保留主值；旧兼容矩阵会展开显示",
                            IssueSeverity.INFO,
                            page_number=page_number,
                            location=worksheet.title,
                        )
                    )
                cells: list[ParsedCell] = []
                for row in range(1, max_row + 1):
                    for column in range(1, max_column + 1):
                        if (row, column) in merged_members:
                            continue
                        cell = worksheet.cell(row, column)
                        value_cell = value_sheet.cell(row, column)
                        formula_text = (
                            str(cell.value)
                            if cell.data_type == "f" and cell.value is not None
                            else None
                        )
                        cached_value = value_cell.value if formula_text else None
                        parsed_value = None
                        raw_value = cached_value if formula_text else cell.value
                        if not formula_text and isinstance(cell.value, date | datetime):
                            parsed_value = cell.value
                            raw_value = to_excel(cell.value, formula_book.epoch)
                        if formula_text and cached_value is None:
                            issues.append(
                                ParseIssue(
                                    "PARSE.FORMULA_VALUE_MISSING",
                                    f"公式{worksheet.title}!{cell.coordinate}没有缓存值，未执行公式",
                                    IssueSeverity.WARNING,
                                    page_number=page_number,
                                    location=f"{worksheet.title}!{cell.coordinate}",
                                )
                            )
                        row_span, column_span = merged_by_master.get((row, column), (1, 1))
                        display_value = cached_value if formula_text else parsed_value or raw_value
                        cells.append(
                            ParsedCell(
                                row_index=row - 1,
                                column_index=column - 1,
                                address=cell.coordinate,
                                raw_value=raw_value,
                                display_text=_display(display_value),
                                value_type="FORMULA"
                                if formula_text
                                else _value_type(parsed_value or raw_value),
                                source_location=f"sheet:{worksheet.title};cell:{cell.coordinate}",
                                extraction_method="XLSX_NATIVE",
                                confidence=1.0,
                                row_span=row_span,
                                column_span=column_span,
                                formula_text=formula_text,
                                cached_value=cached_value,
                                parsed_value=parsed_value,
                                number_format=cell.number_format,
                                bbox=BoundingBox(column - 1, row - 1, column_span, row_span),
                                coordinate_space=CoordinateSpace.WORKSHEET_GRID,
                                source_fragment_sha256=source_fragment_sha256(
                                    [
                                        worksheet.title,
                                        cell.coordinate,
                                        raw_value,
                                        formula_text,
                                        cached_value,
                                    ]
                                ),
                            )
                        )
                hidden_rows = sorted(
                    index
                    for index, dimension in worksheet.row_dimensions.items()
                    if dimension.hidden
                )
                hidden_columns = sorted(
                    name
                    for name, dimension in worksheet.column_dimensions.items()
                    if dimension.hidden
                )
                table = ParsedTable(
                    table_id=f"sheet-{page_number}",
                    row_count=max_row,
                    column_count=max_column,
                    cells=tuple(cells),
                    extraction_method="XLSX_NATIVE",
                    confidence=1.0,
                    page_number=page_number,
                    sheet_name=worksheet.title,
                    coordinate_space=CoordinateSpace.WORKSHEET_GRID,
                    metadata={
                        "legacy_name": worksheet.title,
                        "sheet_state": worksheet.sheet_state,
                        "hidden_rows": hidden_rows,
                        "hidden_columns": hidden_columns,
                        "merged_ranges": [str(item) for item in worksheet.merged_cells.ranges],
                        "freeze_panes": str(worksheet.freeze_panes)
                        if worksheet.freeze_panes
                        else None,
                        "table_ranges": sorted(str(item.ref) for item in worksheet.tables.values()),
                    },
                )
                tables.append(table)
                pages.append(
                    ParsedPage(
                        page_number=page_number,
                        width=float(max_column),
                        height=float(max_row),
                        coordinate_space=CoordinateSpace.WORKSHEET_GRID,
                        classification="STRUCTURED",
                        table_ids=(table.table_id,),
                        metadata={
                            "sheet_name": worksheet.title,
                            "sheet_state": worksheet.sheet_state,
                        },
                    )
                )
        finally:
            formula_book.close()
            value_book.close()
        if not any(table.cells for table in tables):
            issues.append(
                ParseIssue("PARSE.NO_CONTENT", "XLSX没有可解析单元格", IssueSeverity.ERROR)
            )
        return ReaderOutput(
            ParsedDocument(
                document_id=context.source.source_id,
                source=context.source,
                media_type=context.media_type,
                document_kind="XLSX",
                parser_id=self.reader_id,
                parser_version=self.parser_version,
                page_count=len(pages),
                pages=tuple(pages),
                tables=tuple(tables),
                metadata=metadata,
                issues=tuple(issues),
            )
        )

    def supports(self, path: Path) -> bool:
        return path.suffix.casefold() == ".xlsx"

    def read(self, path: Path):
        return legacy_read(
            self,
            path,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class XlsReader:
    reader_id = "xls/xlrd"
    parser_version = "2.0.0"
    media_types = frozenset({"application/vnd.ms-excel"})

    def parse(self, path: Path, context: ParseContext) -> ReaderOutput:
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("读取XLS需要安装xlrd>=2.0") from exc
        try:
            workbook = xlrd.open_workbook(path, on_demand=True, formatting_info=True)
        except (OSError, xlrd.XLRDError) as exc:
            raise ValueError(f"XLS结构无效：{path.name}") from exc
        tables: list[ParsedTable] = []
        pages: list[ParsedPage] = []
        issues: list[ParseIssue] = []
        try:
            for page_number, worksheet in enumerate(workbook.sheets(), start=1):
                if context.cancel_check:
                    context.cancel_check()
                if context.page_progress:
                    context.page_progress(
                        page_number, workbook.nsheets, f"正在解析工作表{worksheet.name}"
                    )
                merged_by_master: dict[tuple[int, int], tuple[int, int]] = {}
                formula_texts, unavailable_formulas = _xls_formula_texts(workbook, page_number - 1)
                if unavailable_formulas:
                    issues.append(
                        ParseIssue(
                            "PARSE.FORMULA_TEXT_UNAVAILABLE",
                            f"工作表{worksheet.name}有{unavailable_formulas}个BIFF公式无法可靠反编译",
                            IssueSeverity.WARNING,
                            page_number=page_number,
                        )
                    )
                merged_members: set[tuple[int, int]] = set()
                for row_low, row_high, col_low, col_high in worksheet.merged_cells:
                    merged_by_master[(row_low, col_low)] = (row_high - row_low, col_high - col_low)
                    for row in range(row_low, row_high):
                        for column in range(col_low, col_high):
                            if (row, column) != (row_low, col_low):
                                merged_members.add((row, column))
                if merged_by_master:
                    issues.append(
                        ParseIssue(
                            "PARSE.MERGED_CELL_EXPANDED",
                            f"工作表{worksheet.name}的合并单元格仅保留主值；旧兼容矩阵会展开显示",
                            IssueSeverity.INFO,
                            page_number=page_number,
                        )
                    )
                cells: list[ParsedCell] = []
                for row in range(worksheet.nrows):
                    for column in range(worksheet.ncols):
                        if (row, column) in merged_members:
                            continue
                        source_cell = worksheet.cell(row, column)
                        raw_value = source_cell.value
                        parsed_value = None
                        if source_cell.ctype == xlrd.XL_CELL_DATE:
                            parsed_value = xlrd.xldate.xldate_as_datetime(
                                raw_value, workbook.datemode
                            )
                        elif source_cell.ctype == xlrd.XL_CELL_BOOLEAN:
                            parsed_value = bool(raw_value)
                        elif source_cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                            raw_value = None
                        row_span, column_span = merged_by_master.get((row, column), (1, 1))
                        formula_text = formula_texts.get((row, column))
                        cached_value = (
                            (parsed_value if parsed_value is not None else raw_value)
                            if formula_text
                            else None
                        )
                        address = f"R{row + 1}C{column + 1}"
                        cells.append(
                            ParsedCell(
                                row_index=row,
                                column_index=column,
                                address=address,
                                raw_value=raw_value,
                                display_text=_display(
                                    parsed_value if parsed_value is not None else raw_value
                                ),
                                value_type=(
                                    "FORMULA"
                                    if formula_text
                                    else _value_type(
                                        parsed_value if parsed_value is not None else raw_value
                                    )
                                ),
                                source_location=f"sheet:{worksheet.name};cell:{address}",
                                extraction_method="XLS_NATIVE",
                                confidence=1.0,
                                row_span=row_span,
                                column_span=column_span,
                                formula_text=formula_text,
                                cached_value=cached_value,
                                parsed_value=parsed_value,
                                bbox=BoundingBox(column, row, column_span, row_span),
                                coordinate_space=CoordinateSpace.WORKSHEET_GRID,
                                source_fragment_sha256=source_fragment_sha256(
                                    [worksheet.name, row, column, raw_value]
                                ),
                            )
                        )
                table = ParsedTable(
                    table_id=f"sheet-{page_number}",
                    row_count=worksheet.nrows,
                    column_count=worksheet.ncols,
                    cells=tuple(cells),
                    extraction_method="XLS_NATIVE",
                    confidence=1.0,
                    page_number=page_number,
                    sheet_name=worksheet.name,
                    coordinate_space=CoordinateSpace.WORKSHEET_GRID,
                    metadata={
                        "legacy_name": worksheet.name,
                        "visibility": worksheet.visibility,
                        "hidden_rows": sorted(
                            index + 1
                            for index, info in worksheet.rowinfo_map.items()
                            if getattr(info, "hidden", False)
                        ),
                        "hidden_columns": sorted(
                            index + 1
                            for index, info in worksheet.colinfo_map.items()
                            if getattr(info, "hidden", False)
                        ),
                    },
                )
                tables.append(table)
                pages.append(
                    ParsedPage(
                        page_number=page_number,
                        width=float(worksheet.ncols),
                        height=float(worksheet.nrows),
                        coordinate_space=CoordinateSpace.WORKSHEET_GRID,
                        classification="STRUCTURED",
                        table_ids=(table.table_id,),
                        metadata={"sheet_name": worksheet.name},
                    )
                )
        finally:
            workbook.release_resources()
        if not any(table.cells for table in tables):
            issues.append(
                ParseIssue("PARSE.NO_CONTENT", "XLS没有可解析单元格", IssueSeverity.ERROR)
            )
        return ReaderOutput(
            ParsedDocument(
                document_id=context.source.source_id,
                source=context.source,
                media_type=context.media_type,
                document_kind="XLS",
                parser_id=self.reader_id,
                parser_version=self.parser_version,
                page_count=len(pages),
                pages=tuple(pages),
                tables=tuple(tables),
                issues=tuple(issues),
            )
        )

    def supports(self, path: Path) -> bool:
        return path.suffix.casefold() == ".xls"

    def read(self, path: Path):
        return legacy_read(self, path, "application/vnd.ms-excel")


__all__ = ["XlsReader", "XlsxReader"]
