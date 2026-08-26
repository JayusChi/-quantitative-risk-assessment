from __future__ import annotations

import csv
import hashlib
import io
from collections.abc import Iterable, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

from ..contracts import RawRow, RawTable, SourceReference

SUPPORTED_SUFFIXES = frozenset({".csv", ".docx", ".pdf", ".xls", ".xlsx"})


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _source(path: Path, reader_id: str) -> SourceReference:
    checksum = _sha256(path)
    return SourceReference(
        source_id=checksum,
        source_path=path.name,
        reader_id=reader_id,
        checksum_sha256=checksum,
    )


def _trim_rows(rows: Iterable[RawRow]) -> tuple[RawRow, ...]:
    result: list[RawRow] = []
    for row in rows:
        cells = list(row.cells)
        while cells and cells[-1] in (None, ""):
            cells.pop()
        if cells:
            result.append(RawRow(row.row_number, tuple(cells)))
    return tuple(result)


class CsvReader:
    reader_id = "csv/stdlib-v1"

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() == ".csv"

    @staticmethod
    def _decode(path: Path) -> str:
        data = path.read_bytes()
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                return data.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise ValueError(f"CSV编码无法识别（仅支持UTF-8/GB18030）：{path.name}")

    def read(self, path: Path) -> Sequence[RawTable]:
        text = self._decode(path)
        try:
            dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(io.StringIO(text, newline=""), dialect)
        raw_rows: list[RawRow] = []
        previous_end_line = 0
        for row in reader:
            raw_rows.append(RawRow(previous_end_line + 1, tuple(row)))
            previous_end_line = reader.line_num
        rows = _trim_rows(raw_rows)
        return (RawTable(_source(path, self.reader_id), path.stem, rows),)


class XlsxReader:
    reader_id = "xlsx/openpyxl-v1"

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() == ".xlsx"

    def read(self, path: Path) -> Sequence[RawTable]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - installation error path
            raise RuntimeError("读取XLSX需要安装openpyxl>=3.1") from exc

        source = _source(path, self.reader_id)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            tables = []
            for worksheet in workbook.worksheets:
                rows = _trim_rows(
                    RawRow(index, tuple(cell.value for cell in row))
                    for index, row in enumerate(worksheet.iter_rows(), start=1)
                )
                tables.append(RawTable(source, worksheet.title, rows))
            return tuple(tables)
        finally:
            workbook.close()


class XlsReader:
    reader_id = "xls/xlrd-v1"

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() == ".xls"

    def read(self, path: Path) -> Sequence[RawTable]:
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover - installation error path
            raise RuntimeError("读取旧版XLS需要安装xlrd>=2.0") from exc

        source = _source(path, self.reader_id)
        workbook = xlrd.open_workbook(path, on_demand=True)
        tables: list[RawTable] = []
        try:
            for worksheet in workbook.sheets():
                converted_rows: list[RawRow] = []
                for row_index in range(worksheet.nrows):
                    cells: list[Any] = []
                    for column_index in range(worksheet.ncols):
                        cell = worksheet.cell(row_index, column_index)
                        value = cell.value
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            value = xlrd.xldate.xldate_as_datetime(value, workbook.datemode)
                        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                            value = bool(value)
                        elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
                            value = None
                        cells.append(value)
                    converted_rows.append(RawRow(row_index + 1, tuple(cells)))
                tables.append(RawTable(source, worksheet.name, _trim_rows(converted_rows)))
            return tuple(tables)
        finally:
            workbook.release_resources()


class ReaderRegistry:
    def __init__(self) -> None:
        from .documents import DocxReader
        from .pdf import PdfReader

        self._readers = (CsvReader(), DocxReader(), PdfReader(), XlsReader(), XlsxReader())

    @property
    def supported_suffixes(self) -> frozenset[str]:
        return SUPPORTED_SUFFIXES

    def reader_for(self, path: Path) -> Any:
        for reader in self._readers:
            if reader.supports(path):
                return reader
        raise ValueError(f"不支持的源文件格式：{path.suffix or path.name}")

    def read(self, path: Path) -> Sequence[RawTable]:
        return self.reader_for(path).read(path)


def json_safe_cell(value: Any) -> Any:
    if isinstance(value, date | datetime):
        return value.isoformat()
    return value


__all__ = [
    "CsvReader",
    "ReaderRegistry",
    "SUPPORTED_SUFFIXES",
    "XlsReader",
    "XlsxReader",
    "json_safe_cell",
]
