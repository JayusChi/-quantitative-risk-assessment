"""Application service for human review, evidence, gating and confirmation."""

from __future__ import annotations

import csv
import io
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

import pdfplumber
import xlrd
from openpyxl import load_workbook
from PIL import Image, ImageDraw

from qra_converter.contract_catalog import load_contract_catalog

from .database import QraDatabase, canonical_json, json_sha256, utc_now
from .review_assembly import (
    ASSEMBLY_RULE_VERSION,
    ReviewAssemblyError,
    assemble_review_payload,
    field_definitions,
    normalize_manual_value,
)
from .review_gate import GATE_RULE_VERSION, evaluate_review_gate

EDITABLE_SESSION_STATUSES = {"OPEN", "IN_REVIEW", "READY_TO_CONFIRM"}
DECISION_ACTIONS = {
    "ACCEPT_CANDIDATE",
    "OVERRIDE_VALUE",
    "REJECT_ALL",
    "MARK_NOT_APPLICABLE",
    "REQUEST_REEXTRACTION",
}
REVIEW_SERVICE_VERSION = "qra.review-service/1.0.0"


class ReviewRevisionConflict(ValueError):
    """The client attempted to overwrite a newer review revision."""

    code = "REVIEW_REVISION_CONFLICT"

    def __init__(self, session: dict[str, Any]):
        super().__init__("复核会话已被其他操作更新，请刷新后重试")
        self.session = session


def _json(value: Any, fallback: Any) -> Any:
    if value is None:
        return fallback
    if isinstance(value, dict | list):
        return value
    try:
        return json.loads(str(value))
    except (TypeError, json.JSONDecodeError):
        return fallback


def _highlight_bbox(
    image: Image.Image,
    bbox: object,
    source_size: object,
) -> Image.Image:
    """Draw an xywh evidence box after scaling source coordinates to preview pixels."""

    if not isinstance(bbox, list | tuple) or len(bbox) != 4:
        return image
    try:
        x, y, width, height = (float(value) for value in bbox)
        if isinstance(source_size, list | tuple) and len(source_size) == 2:
            source_width, source_height = (max(1.0, float(value)) for value in source_size)
        else:
            source_width, source_height = float(image.width), float(image.height)
    except (TypeError, ValueError):
        return image
    scale_x, scale_y = image.width / source_width, image.height / source_height
    x0, y0 = max(0.0, x * scale_x), max(0.0, y * scale_y)
    x1 = min(float(image.width), (x + max(0.0, width)) * scale_x)
    y1 = min(float(image.height), (y + max(0.0, height)) * scale_y)
    if x1 <= x0 or y1 <= y0:
        return image
    highlighted = image.convert("RGBA")
    overlay = Image.new("RGBA", highlighted.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    line_width = max(2, min(highlighted.size) // 250)
    draw.rectangle(
        (x0, y0, x1, y1),
        fill=(239, 68, 68, 45),
        outline=(220, 38, 38, 255),
        width=line_width,
    )
    return Image.alpha_composite(highlighted, overlay)


def _png_bytes(image: Image.Image) -> bytes:
    output = io.BytesIO()
    image.save(output, format="PNG", optimize=True)
    return output.getvalue()


class ReviewService:
    def __init__(self, database: QraDatabase):
        self.database = database
        self.database.initialize()

    @staticmethod
    def _session_document(row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
        item = dict(row)
        item["target_node_ids"] = _json(item.pop("target_node_ids_json", None), [])
        return item

    @staticmethod
    def _job_row(connection: sqlite3.Connection, job_id: str) -> sqlite3.Row:
        row = connection.execute("SELECT * FROM conversion_job WHERE id = ?", (job_id,)).fetchone()
        if row is None:
            raise KeyError(f"转换任务不存在：{job_id}")
        return row

    @staticmethod
    def _candidate_set_hash(connection: sqlite3.Connection, job: sqlite3.Row) -> str:
        rows = connection.execute(
            """
            SELECT candidate_id, payload_json FROM candidate_field
            WHERE job_id = ? ORDER BY candidate_id
            """,
            (job["id"],),
        ).fetchall()
        candidates = []
        for row in rows:
            candidate = _json(row["payload_json"], {})
            evidence = [
                _json(link["evidence_json"], {})
                for link in connection.execute(
                    """
                    SELECT evidence_json FROM candidate_evidence_link
                    WHERE job_id = ? AND candidate_id = ? ORDER BY evidence_id
                    """,
                    (job["id"], row["candidate_id"]),
                ).fetchall()
            ]
            candidates.append({"candidate": candidate, "evidence": evidence})
        return json_sha256(
            {
                "candidates": candidates,
                "contract": {
                    "id": job["contract_id"],
                    "version": job["contract_version"],
                    "sha256": job["contract_sha256"],
                },
                "mapping": {
                    "id": job["profile_id"],
                    "version": job["profile_version"],
                    "sha256": job["profile_sha256"],
                },
                "stage4_result_sha256": job["stage4_result_sha256"],
            }
        )

    @staticmethod
    def _source_manifest_hash(job: sqlite3.Row) -> str:
        manifest = _json(job["source_manifest_json"], {})
        return json_sha256(manifest) if manifest else str(job["file_manifest_sha256"] or "")

    @staticmethod
    def _catalog(job: sqlite3.Row) -> Any:
        return load_contract_catalog(
            Path(str(job["contract_path"])),
            expected_contract_id=str(job["contract_id"]),
            expected_version=str(job["contract_version"]),
            expected_manifest_sha256=str(job["contract_sha256"]),
        )

    @staticmethod
    def _unit_registry(catalog: Any) -> dict[str, Any]:
        relative = str(catalog.manifest.get("unit_registry") or "unit_registry.json")
        return json.loads((catalog.root / relative).read_text(encoding="utf-8"))

    @staticmethod
    def _default_targets(job: sqlite3.Row) -> list[str]:
        capability = _json(job["stage4_capability_json"], {})
        plan = capability.get("engine_import_preflight", {}).get("dynamic_plan", {})
        targets = [str(value) for value in plan.get("runnable_node_ids", []) if value]
        if "segment_geometry" not in targets:
            targets.insert(0, "segment_geometry")
        return list(dict.fromkeys(targets))

    def create_or_resume_session(
        self,
        job_id: str,
        *,
        actor: str,
        target_node_ids: list[str] | None = None,
        owner: str | None = None,
    ) -> tuple[dict[str, Any], bool]:
        safe_actor = str(actor).strip()[:120]
        if not safe_actor:
            raise ValueError("复核会话创建人不能为空")
        with self.database.transaction() as connection:
            job = self._job_row(connection, job_id)
            if str(job["status"]) not in {"BLOCKED", "READY_FOR_CONFIRMATION", "CONFIRMED"}:
                raise ValueError("只有已完成转换的任务可以进入复核工作台")
            candidate_hash = self._candidate_set_hash(connection, job)
            if not connection.execute(
                "SELECT 1 FROM candidate_field WHERE job_id = ? LIMIT 1", (job_id,)
            ).fetchone():
                raise ValueError("转换任务尚未产生候选字段")
            current = connection.execute(
                """
                SELECT * FROM review_session
                WHERE conversion_job_id = ? AND status IN ('OPEN','IN_REVIEW','READY_TO_CONFIRM')
                ORDER BY created_at DESC LIMIT 1
                """,
                (job_id,),
            ).fetchone()
            if current is not None and str(current["candidate_set_hash"]) == candidate_hash:
                current_doc = self._session_document(current)
                requested = target_node_ids or current_doc["target_node_ids"]
                if list(requested) != list(current_doc["target_node_ids"]):
                    raise ValueError("已有活动会话的目标节点范围不同；请先完成或废弃该会话")
                self.database._record_event_in_connection(
                    connection,
                    event_type="REVIEW_SESSION_RESUMED",
                    entity_type="review_session",
                    entity_id=str(current["id"]),
                    actor=safe_actor,
                    detail={"conversion_job_id": job_id, "revision": current["revision"]},
                )
                return current_doc, False
            now = utc_now()
            session_id = f"RSESS-{uuid4()}"
            if current is not None:
                connection.execute(
                    """
                    UPDATE review_session
                    SET status = 'SUPERSEDED', superseded_by_session_id = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (session_id, now, current["id"]),
                )
                self.database._record_event_in_connection(
                    connection,
                    event_type="REVIEW_SESSION_STALE",
                    entity_type="review_session",
                    entity_id=str(current["id"]),
                    actor=safe_actor,
                    detail={"conversion_job_id": job_id, "new_candidate_set_hash": candidate_hash},
                )
            targets = [str(value) for value in (target_node_ids or self._default_targets(job))]
            if not targets:
                raise ValueError("复核会话必须声明至少一个目标计算节点")
            connection.execute(
                """
                INSERT INTO review_session(
                    id, conversion_job_id, status, revision, candidate_set_hash,
                    source_manifest_hash, target_node_ids_json, owner, created_by,
                    created_at, updated_at
                ) VALUES (?, ?, 'OPEN', 1, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    session_id,
                    job_id,
                    candidate_hash,
                    self._source_manifest_hash(job),
                    canonical_json(targets),
                    (owner or safe_actor)[:120],
                    safe_actor,
                    now,
                    now,
                ),
            )
            self.database._record_event_in_connection(
                connection,
                event_type="REVIEW_SESSION_CREATED",
                entity_type="review_session",
                entity_id=session_id,
                actor=safe_actor,
                detail={
                    "conversion_job_id": job_id,
                    "candidate_set_hash": candidate_hash,
                    "target_node_ids": targets,
                    "revision": 1,
                },
            )
            row = connection.execute(
                "SELECT * FROM review_session WHERE id = ?", (session_id,)
            ).fetchone()
        return self._session_document(row), True

    @staticmethod
    def _effective_decisions(
        connection: sqlite3.Connection, session_id: str
    ) -> list[dict[str, Any]]:
        rows = connection.execute(
            """
            SELECT d.* FROM review_decision d
            WHERE d.session_id = ?
              AND NOT EXISTS (
                  SELECT 1 FROM review_decision newer
                  WHERE newer.supersedes_decision_id = d.id
              )
            ORDER BY d.review_item_key
            """,
            (session_id,),
        ).fetchall()
        result = []
        for row in rows:
            item = dict(row)
            item["override_raw_value"] = _json(item.pop("override_raw_value_json"), None)
            item["override_normalized_value"] = _json(
                item.pop("override_normalized_value_json"), None
            )
            result.append(item)
        return result

    @staticmethod
    def _decision_set_hash(decisions: list[dict[str, Any]]) -> str:
        stable = [
            {
                "review_item_key": row["review_item_key"],
                "field_id": row["field_id"],
                "entity_id": row.get("entity_id"),
                "action": row["action"],
                "selected_candidate_id": row.get("selected_candidate_id"),
                "override_raw_value": row.get("override_raw_value"),
                "override_normalized_value": row.get("override_normalized_value"),
                "override_unit": row.get("override_unit"),
                "applicability_reason": row.get("applicability_reason"),
                "reason": row.get("reason"),
            }
            for row in sorted(decisions, key=lambda value: str(value["review_item_key"]))
        ]
        return json_sha256(stable)

    @staticmethod
    def _review_item_key(entity_key: str, field_id: str) -> str:
        return f"RITEM-{json_sha256({'entity_key': entity_key, 'field_id': field_id})[:24]}"

    def _build_items(
        self,
        connection: sqlite3.Connection,
        *,
        session: dict[str, Any],
        definitions: dict[str, dict[str, Any]],
    ) -> list[dict[str, Any]]:
        job_id = str(session["conversion_job_id"])
        source_rows = connection.execute(
            """
            SELECT id, relative_path, original_file_name, detected_media_type, sha256,
                   parser_id, parser_version, parse_quality_json, parsed_at
            FROM conversion_source WHERE job_id = ?
            """,
            (job_id,),
        ).fetchall()
        sources = {str(row["id"]): dict(row) for row in source_rows}
        grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
        for row in connection.execute(
            "SELECT payload_json FROM candidate_field WHERE job_id = ? ORDER BY candidate_id",
            (job_id,),
        ).fetchall():
            candidate = _json(row["payload_json"], {})
            entity = candidate.get("entity") or {}
            key = (str(entity.get("entity_key") or "GLOBAL"), str(candidate.get("field_id")))
            evidence_rows = connection.execute(
                """
                SELECT evidence_json FROM candidate_evidence_link
                WHERE job_id = ? AND candidate_id = ? ORDER BY evidence_id
                """,
                (job_id, candidate.get("candidate_id")),
            ).fetchall()
            candidate["evidence"] = []
            for evidence_row in evidence_rows:
                evidence = _json(evidence_row["evidence_json"], {})
                source_id = str((evidence.get("location") or {}).get("file_id") or "")
                source = sources.get(source_id)
                if source:
                    evidence["source"] = {
                        "source_id": source_id,
                        "file_name": source.get("relative_path")
                        or source.get("original_file_name"),
                        "media_type": source.get("detected_media_type"),
                        "sha256": source.get("sha256"),
                    }
                candidate["evidence"].append(evidence)
            grouped.setdefault(key, []).append(candidate)

        issues = [
            _json(row["payload_json"], {})
            for row in connection.execute(
                "SELECT payload_json FROM quality_issue WHERE job_id = ? ORDER BY issue_id",
                (job_id,),
            ).fetchall()
        ]
        for issue in issues:
            field_id = str(issue.get("field_id") or "")
            issue_key = (str(issue.get("entity_key") or "GLOBAL"), field_id)
            if field_id and issue_key not in grouped:
                grouped[issue_key] = []

        fusion: dict[tuple[str, str], list[dict[str, Any]]] = {}
        for row in connection.execute(
            """
            SELECT entity_key, field_id, payload_json FROM fusion_group
            WHERE job_id = ? AND field_id IS NOT NULL ORDER BY fusion_group_id
            """,
            (job_id,),
        ).fetchall():
            fusion.setdefault((str(row["entity_key"]), str(row["field_id"])), []).append(
                _json(row["payload_json"], {})
            )
        decisions = {
            str(row["review_item_key"]): row
            for row in self._effective_decisions(connection, str(session["id"]))
        }
        target_nodes = set(session.get("target_node_ids") or [])
        items = []
        for (entity_key, field_id), candidates in grouped.items():
            definition = definitions.get(field_id, {})
            review_key = self._review_item_key(entity_key, field_id)
            candidate_ids = {str(row.get("candidate_id")) for row in candidates}
            item_issues = [
                issue
                for issue in issues
                if str(issue.get("field_id") or "") == field_id
                and str(issue.get("entity_key") or entity_key) == entity_key
                and (
                    not issue.get("candidate_ids")
                    or bool(
                        candidate_ids & {str(value) for value in issue.get("candidate_ids", [])}
                    )
                )
            ]
            groups = fusion.get((entity_key, field_id), [])
            distinct_values = {canonical_json(row.get("normalized_value")) for row in candidates}
            conflict = len(distinct_values) > 1 or any(
                str(group.get("group_type")) in {"CONFLICT", "IDENTITY_AMBIGUOUS"}
                for group in groups
            )
            low_confidence = any(
                str(row.get("quality_status")) == "LOW_CONFIDENCE"
                or float(row.get("confidence", 0.0)) < 0.8
                for row in candidates
            )
            invalid = any(str(row.get("quality_status")) == "INVALID" for row in candidates)
            required_nodes = [str(value) for value in definition.get("required_by_nodes", ())]
            required = str(definition.get("required_level")) == "REQUIRED" or bool(
                target_nodes & set(required_nodes)
            )
            missing = not candidates
            blocking_issue = any(bool(issue.get("blocking")) for issue in item_issues)
            base_requires = (
                conflict or invalid or blocking_issue or (required and (missing or low_confidence))
            )
            deterministic = bool(
                len(candidates) == 1
                and not conflict
                and not invalid
                and str(candidates[0].get("quality_status")) == "PASS"
                and candidates[0].get("evidence")
            )
            decision = decisions.get(review_key)
            action = str(decision.get("action")) if decision else None
            status = {
                "ACCEPT_CANDIDATE": "ACCEPTED",
                "OVERRIDE_VALUE": "OVERRIDDEN",
                "REJECT_ALL": "REJECTED",
                "MARK_NOT_APPLICABLE": "NOT_APPLICABLE",
                "REQUEST_REEXTRACTION": "REEXTRACTION_REQUESTED",
            }.get(action, "AUTO_DETERMINISTIC" if deterministic else "UNRESOLVED")
            resolved = action in {"ACCEPT_CANDIDATE", "OVERRIDE_VALUE", "MARK_NOT_APPLICABLE"}
            requires_resolution = bool(
                (base_requires and not resolved) or action in {"REJECT_ALL", "REQUEST_REEXTRACTION"}
            )
            proposed = next(
                (
                    group.get("proposed_candidate_id")
                    for group in groups
                    if group.get("proposed_candidate_id")
                ),
                candidates[0].get("candidate_id") if len(candidates) == 1 else None,
            )
            item = {
                "review_item_key": review_key,
                "entity_id": entity_key,
                "entity_type": candidates[0].get("entity", {}).get("entity_type")
                if candidates
                else definition.get("entity_type"),
                "field_id": field_id,
                "field_name": definition.get("name_zh") or field_id,
                "field_group": field_id.split(".", 1)[0],
                "field_definition": definition,
                "candidates": candidates,
                "candidate_count": len(candidates),
                "evidence_count": sum(len(row.get("evidence", [])) for row in candidates),
                "highest_confidence": max(
                    (float(row.get("confidence", 0.0)) for row in candidates), default=0.0
                ),
                "fusion_groups": groups,
                "quality_issues": item_issues,
                "issue_codes": sorted({str(issue.get("code")) for issue in item_issues}),
                "conflict": conflict,
                "low_confidence": low_confidence,
                "missing": missing,
                "invalid": invalid,
                "blocking": base_requires,
                "required": required,
                "affected_node_ids": required_nodes,
                "proposed_candidate_id": proposed,
                "current_decision": decision,
                "resolution_status": status,
                "requires_resolution": requires_resolution,
                "available_actions": [
                    "ACCEPT_CANDIDATE",
                    "OVERRIDE_VALUE",
                    "REJECT_ALL",
                    *(
                        ["MARK_NOT_APPLICABLE"]
                        if str(definition.get("required_level")) == "OPTIONAL"
                        else []
                    ),
                    "REQUEST_REEXTRACTION",
                ],
            }
            items.append(item)
        return sorted(
            items,
            key=lambda item: (
                not item["requires_resolution"],
                not item["blocking"],
                -len(item["affected_node_ids"]),
                not item["conflict"],
                item["highest_confidence"],
                item["field_name"],
                item["entity_id"],
            ),
        )

    def _load_session_context(
        self, connection: sqlite3.Connection, session_id: str
    ) -> tuple[dict[str, Any], sqlite3.Row, Any, dict[str, dict[str, Any]]]:
        row = connection.execute(
            "SELECT * FROM review_session WHERE id = ?", (session_id,)
        ).fetchone()
        if row is None:
            raise KeyError(f"复核会话不存在：{session_id}")
        session = self._session_document(row)
        job = self._job_row(connection, str(session["conversion_job_id"]))
        catalog = self._catalog(job)
        return session, job, catalog, field_definitions(catalog)

    def get_session(self, session_id: str) -> dict[str, Any]:
        with self.database.session() as connection:
            session, job, catalog, definitions = self._load_session_context(connection, session_id)
            current_hash = self._candidate_set_hash(connection, job)
            if (
                session["status"] in EDITABLE_SESSION_STATUSES
                and current_hash != session["candidate_set_hash"]
            ):
                connection.execute(
                    "UPDATE review_session SET status = 'STALE', updated_at = ? WHERE id = ?",
                    (utc_now(), session_id),
                )
                self.database._record_event_in_connection(
                    connection,
                    event_type="REVIEW_SESSION_STALE",
                    entity_type="review_session",
                    entity_id=session_id,
                    detail={
                        "conversion_job_id": session["conversion_job_id"],
                        "old_candidate_set_hash": session["candidate_set_hash"],
                        "new_candidate_set_hash": current_hash,
                    },
                )
                session["status"] = "STALE"
            decisions = self._effective_decisions(connection, session_id)
            items = self._build_items(connection, session=session, definitions=definitions)
            latest_gate = connection.execute(
                """
                SELECT result_json FROM review_gate_run
                WHERE session_id = ? ORDER BY created_at DESC, id DESC LIMIT 1
                """,
                (session_id,),
            ).fetchone()
        resolved = sum(1 for item in items if not item["requires_resolution"])
        session.update(
            {
                "project_name": job["project_name"] or job["case_id"],
                "conversion_status": job["status"],
                "decision_set_hash": self._decision_set_hash(decisions),
                "progress": {
                    "total": len(items),
                    "resolved": resolved,
                    "unresolved": len(items) - resolved,
                    "percent": round(resolved / len(items) * 100) if items else 100,
                },
                "gate_summary": _json(latest_gate["result_json"], None) if latest_gate else None,
                "service_version": REVIEW_SERVICE_VERSION,
                "contract_version": catalog.version,
            }
        )
        return session

    def list_items(
        self,
        session_id: str,
        *,
        status: str | None = None,
        severity: str | None = None,
        blocking: bool | None = None,
        field_group: str | None = None,
        source_id: str | None = None,
        node_id: str | None = None,
        q: str | None = None,
        offset: int = 0,
        limit: int = 100,
    ) -> dict[str, Any]:
        session = self.get_session(session_id)
        with self.database.session() as connection:
            _, _, _, definitions = self._load_session_context(connection, session_id)
            items = self._build_items(connection, session=session, definitions=definitions)
        query = str(q or "").strip().casefold()
        filtered = []
        for item in items:
            if status and item["resolution_status"] != status:
                continue
            if severity and not any(
                str(issue.get("quality_status") or issue.get("severity")) == severity
                for issue in item["quality_issues"]
            ):
                continue
            if blocking is not None and bool(item["blocking"]) != blocking:
                continue
            if field_group and item["field_group"] != field_group:
                continue
            if node_id and node_id not in item["affected_node_ids"]:
                continue
            if source_id and not any(
                str((evidence.get("location") or {}).get("file_id")) == source_id
                for candidate in item["candidates"]
                for evidence in candidate.get("evidence", [])
            ):
                continue
            haystack = f"{item['field_id']} {item['field_name']} {item['entity_id']}".casefold()
            if query and query not in haystack:
                continue
            filtered.append(item)
        start = max(0, int(offset))
        safe_limit = max(1, min(int(limit), 500))
        return {
            "items": filtered[start : start + safe_limit],
            "total": len(filtered),
            "offset": start,
            "limit": safe_limit,
        }

    def get_item(self, session_id: str, review_item_key: str) -> dict[str, Any]:
        result = self.list_items(session_id, limit=500)
        for item in result["items"]:
            if item["review_item_key"] == review_item_key:
                return item
        raise KeyError(f"复核项不存在：{review_item_key}")

    def save_decision(
        self,
        session_id: str,
        *,
        review_item_key: str,
        action: str,
        selected_candidate_id: str | None,
        override_value: Any,
        override_unit: str | None,
        reason: str,
        actor: str,
        expected_revision: int,
        source_id: str | None = None,
        evidence_id: str | None = None,
    ) -> dict[str, Any]:
        action = str(action).strip().upper()
        if action not in DECISION_ACTIONS:
            raise ValueError(f"不支持的复核操作：{action}")
        safe_reason = str(reason).strip()[:1000]
        safe_actor = str(actor).strip()[:120]
        if not safe_reason or not safe_actor:
            raise ValueError("复核操作必须填写确认人和原因")
        with self.database.transaction() as connection:
            session, job, catalog, definitions = self._load_session_context(connection, session_id)
            if session["status"] not in EDITABLE_SESSION_STATUSES:
                raise ValueError("复核会话当前不可编辑")
            if int(session["revision"]) != int(expected_revision):
                raise ReviewRevisionConflict(self.get_session(session_id))
            current_hash = self._candidate_set_hash(connection, job)
            if current_hash != session["candidate_set_hash"]:
                raise ValueError("候选集合已经变化，旧会话必须作废并重新复核")
            items = self._build_items(connection, session=session, definitions=definitions)
            item = next((row for row in items if row["review_item_key"] == review_item_key), None)
            if item is None:
                raise KeyError(f"复核项不存在：{review_item_key}")
            definition = definitions.get(str(item["field_id"]))
            if definition is None:
                raise ValueError("复核字段不在当前合同中")
            normalized_value = None
            raw_value = None
            selected = None
            applicability_reason = None
            candidate_ids = {str(row["candidate_id"]) for row in item["candidates"]}
            if action == "ACCEPT_CANDIDATE":
                selected = str(selected_candidate_id or "")
                if selected not in candidate_ids:
                    raise ValueError("选择的候选不属于当前复核项或已经过期")
                candidate = next(
                    row for row in item["candidates"] if row["candidate_id"] == selected
                )
                if str(candidate.get("quality_status")) == "INVALID":
                    raise ValueError("无效候选不能接受，请手工修正或请求重新提取")
                if not candidate.get("evidence"):
                    raise ValueError("文档候选没有有效证据，不能接受")
            elif action == "OVERRIDE_VALUE":
                raw_value = override_value
                normalized_value = normalize_manual_value(
                    field_id=str(item["field_id"]),
                    value=override_value,
                    unit=str(override_unit).strip() if override_unit else None,
                    definition=definition,
                    unit_registry=self._unit_registry(catalog),
                )
            elif action == "MARK_NOT_APPLICABLE":
                if str(definition.get("required_level")) != "OPTIONAL":
                    raise ValueError("该字段合同不允许标记为不适用")
                applicability_reason = safe_reason
            elif action == "REQUEST_REEXTRACTION":
                if evidence_id and not any(
                    str(evidence.get("evidence_id")) == str(evidence_id)
                    for candidate in item["candidates"]
                    for evidence in candidate.get("evidence", [])
                ):
                    raise ValueError("重提取证据不属于当前复核项")
                if source_id and source_id not in {
                    str((evidence.get("location") or {}).get("file_id"))
                    for candidate in item["candidates"]
                    for evidence in candidate.get("evidence", [])
                }:
                    raise ValueError("重提取源文件不属于当前复核项")
            previous = next(
                (
                    row
                    for row in self._effective_decisions(connection, session_id)
                    if row["review_item_key"] == review_item_key
                ),
                None,
            )
            new_revision = int(session["revision"]) + 1
            stable_decision = {
                "review_item_key": review_item_key,
                "entity_id": item["entity_id"],
                "field_id": item["field_id"],
                "action": action,
                "selected_candidate_id": selected,
                "override_raw_value": raw_value,
                "override_normalized_value": normalized_value,
                "override_unit": override_unit,
                "applicability_reason": applicability_reason,
                "reason": safe_reason,
                "candidate_set_hash": current_hash,
            }
            decision_id = f"RDEC-{uuid4()}"
            connection.execute(
                """
                INSERT INTO review_decision(
                    id, session_id, review_item_key, entity_id, field_id, action,
                    selected_candidate_id, override_raw_value_json,
                    override_normalized_value_json, override_unit,
                    applicability_reason, reason, reviewer, session_revision,
                    candidate_set_hash, decision_sha256, created_at,
                    supersedes_decision_id, active
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                """,
                (
                    decision_id,
                    session_id,
                    review_item_key,
                    item["entity_id"],
                    item["field_id"],
                    action,
                    selected,
                    canonical_json(raw_value) if action == "OVERRIDE_VALUE" else None,
                    canonical_json(normalized_value) if action == "OVERRIDE_VALUE" else None,
                    override_unit,
                    applicability_reason,
                    safe_reason,
                    safe_actor,
                    new_revision,
                    current_hash,
                    json_sha256(stable_decision),
                    utc_now(),
                    previous["id"] if previous else None,
                ),
            )
            request_id = None
            if action == "REQUEST_REEXTRACTION":
                request_id = f"REXT-{uuid4()}"
                connection.execute(
                    """
                    INSERT INTO reextraction_request(
                        id, session_id, conversion_job_id, source_id, field_id,
                        entity_id, evidence_id, requested_by, reason, status, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'QUEUED', ?)
                    """,
                    (
                        request_id,
                        session_id,
                        session["conversion_job_id"],
                        source_id,
                        item["field_id"],
                        item["entity_id"],
                        evidence_id,
                        safe_actor,
                        safe_reason,
                        utc_now(),
                    ),
                )
            connection.execute(
                """
                UPDATE review_session
                SET status = 'IN_REVIEW', revision = ?, updated_at = ?,
                    gate_status = NULL, gate_result_hash = NULL, decision_set_hash = NULL
                WHERE id = ?
                """,
                (new_revision, utc_now(), session_id),
            )
            event_type = {
                "ACCEPT_CANDIDATE": "REVIEW_DECISION_ACCEPTED",
                "OVERRIDE_VALUE": "REVIEW_DECISION_OVERRIDDEN",
                "REJECT_ALL": "REVIEW_DECISION_REJECTED",
                "MARK_NOT_APPLICABLE": "REVIEW_DECISION_NOT_APPLICABLE",
                "REQUEST_REEXTRACTION": "REVIEW_REEXTRACTION_REQUESTED",
            }[action]
            self.database._record_event_in_connection(
                connection,
                event_type=event_type,
                entity_type="review_session",
                entity_id=session_id,
                actor=safe_actor,
                detail={
                    "conversion_job_id": session["conversion_job_id"],
                    "review_session_id": session_id,
                    "review_item_key": review_item_key,
                    "field_id": item["field_id"],
                    "old_decision_id": previous["id"] if previous else None,
                    "new_decision_id": decision_id,
                    "revision": new_revision,
                    "candidate_set_hash": current_hash,
                },
            )
        updated = self.get_session(session_id)
        return {
            "decision_id": decision_id,
            "reextraction_request_id": request_id,
            "session_revision": updated["revision"],
            "session": updated,
            "item": self.get_item(session_id, review_item_key),
        }

    def complete_reextraction(
        self,
        request_id: str,
        *,
        actor: str,
        replacement_extraction_run_id: str | None = None,
    ) -> dict[str, Any]:
        """Close a queued request; candidate changes invalidate the editable session."""

        safe_actor = str(actor).strip()
        if not safe_actor:
            raise ValueError("完成重新提取必须提供操作人")
        with self.database.transaction() as connection:
            row = connection.execute(
                """
                SELECT request.*, session.candidate_set_hash AS session_candidate_set_hash,
                       session.status AS session_status
                FROM reextraction_request request
                JOIN review_session session ON session.id = request.session_id
                WHERE request.id = ?
                """,
                (request_id,),
            ).fetchone()
            if row is None:
                raise KeyError(f"重新提取请求不存在：{request_id}")
            if str(row["status"]) not in {"QUEUED", "RUNNING"}:
                raise ValueError("重新提取请求已处于终态")
            if replacement_extraction_run_id is not None:
                linked = connection.execute(
                    "SELECT id FROM extraction_run WHERE id = ? AND job_id = ?",
                    (replacement_extraction_run_id, row["conversion_job_id"]),
                ).fetchone()
                if linked is None:
                    raise ValueError("替换提取运行不属于当前转换任务")
            finished_at = utc_now()
            connection.execute(
                """
                UPDATE reextraction_request
                SET status = 'COMPLETED', finished_at = ?, replacement_extraction_run_id = ?
                WHERE id = ?
                """,
                (finished_at, replacement_extraction_run_id, request_id),
            )
            job = self._job_row(connection, str(row["conversion_job_id"]))
            current_hash = self._candidate_set_hash(connection, job)
            became_stale = bool(
                row["session_status"] in EDITABLE_SESSION_STATUSES
                and current_hash != row["session_candidate_set_hash"]
            )
            if became_stale:
                connection.execute(
                    """
                    UPDATE review_session
                    SET status = 'STALE', gate_status = NULL, gate_result_hash = NULL,
                        updated_at = ? WHERE id = ?
                    """,
                    (finished_at, row["session_id"]),
                )
            detail = {
                "conversion_job_id": row["conversion_job_id"],
                "review_session_id": row["session_id"],
                "field_id": row["field_id"],
                "reextraction_request_id": request_id,
                "replacement_extraction_run_id": replacement_extraction_run_id,
                "candidate_set_hash": current_hash,
            }
            self.database._record_event_in_connection(
                connection,
                event_type="REVIEW_REEXTRACTION_COMPLETED",
                entity_type="reextraction_request",
                entity_id=request_id,
                actor=safe_actor,
                detail=detail,
            )
            if became_stale:
                self.database._record_event_in_connection(
                    connection,
                    event_type="REVIEW_SESSION_STALE",
                    entity_type="review_session",
                    entity_id=str(row["session_id"]),
                    actor=safe_actor,
                    detail=detail,
                )
        return {
            "id": request_id,
            "status": "COMPLETED",
            "session_id": row["session_id"],
            "session_stale": became_stale,
            "candidate_set_hash": current_hash,
            "finished_at": finished_at,
        }

    def _evaluate(
        self,
        connection: sqlite3.Connection,
        *,
        session: dict[str, Any],
        job: sqlite3.Row,
        catalog: Any,
        definitions: dict[str, dict[str, Any]],
    ) -> tuple[dict[str, Any], dict[str, Any] | None, list[dict[str, Any]]]:
        decisions = self._effective_decisions(connection, str(session["id"]))
        items = self._build_items(connection, session=session, definitions=definitions)
        candidate_rows = connection.execute(
            "SELECT candidate_id, payload_json FROM candidate_field WHERE job_id = ?",
            (job["id"],),
        ).fetchall()
        candidates = {
            str(row["candidate_id"]): _json(row["payload_json"], {}) for row in candidate_rows
        }
        payload = None
        assembly_error = None
        try:
            base_payload = _json(job["payload_json"], None)
            if not isinstance(base_payload, dict):
                raise ReviewAssemblyError("转换任务没有可组装的JSON草稿")
            payload = assemble_review_payload(
                base_payload=base_payload,
                decisions=decisions,
                candidates=candidates,
                definitions=definitions,
            )
        except (ReviewAssemblyError, TypeError, ValueError) as exc:
            assembly_error = str(exc)
        candidate_hash = self._candidate_set_hash(connection, job)
        decision_hash = self._decision_set_hash(decisions)
        result = evaluate_review_gate(
            session=session,
            items=items,
            decisions=decisions,
            payload=payload,
            assembly_error=assembly_error,
            catalog=catalog,
            candidate_set_hash=candidate_hash,
            decision_set_hash=decision_hash,
        )
        return result, payload, decisions

    def _insert_gate(
        self,
        connection: sqlite3.Connection,
        *,
        session: dict[str, Any],
        result: dict[str, Any],
    ) -> str:
        gate_id = f"RGATE-{uuid4()}"
        connection.execute(
            """
            INSERT INTO review_gate_run(
                id, session_id, session_revision, candidate_set_hash,
                decision_set_hash, status, blocking_issue_count, warning_count,
                unresolved_field_count, accepted_field_count,
                overridden_field_count, rejected_field_count,
                not_applicable_count, runnable_node_ids_json,
                blocked_node_ids_json, missing_inputs_json,
                assembled_payload_hash, result_hash, result_json, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                gate_id,
                session["id"],
                session["revision"],
                result["candidate_set_hash"],
                result["decision_set_hash"],
                result["status"],
                result["blocking_issue_count"],
                result["warning_count"],
                result["unresolved_field_count"],
                result["accepted_field_count"],
                result["overridden_field_count"],
                result["rejected_field_count"],
                result["not_applicable_count"],
                canonical_json(result["runnable_node_ids"]),
                canonical_json(result["blocked_node_ids"]),
                canonical_json(result["missing_inputs"]),
                result["payload_sha256"],
                result["result_hash"],
                canonical_json(result),
                utc_now(),
            ),
        )
        return gate_id

    def run_gate(self, session_id: str, *, actor: str) -> dict[str, Any]:
        safe_actor = str(actor).strip()[:120] or "local-reviewer"
        with self.database.transaction() as connection:
            session, job, catalog, definitions = self._load_session_context(connection, session_id)
            if session["status"] not in EDITABLE_SESSION_STATUSES:
                raise ValueError("复核会话当前不能执行门禁")
            if self._candidate_set_hash(connection, job) != session["candidate_set_hash"]:
                raise ValueError("候选集合已经变化，必须创建新复核会话")
            self.database._record_event_in_connection(
                connection,
                event_type="REVIEW_GATE_STARTED",
                entity_type="review_session",
                entity_id=session_id,
                actor=safe_actor,
                detail={
                    "conversion_job_id": session["conversion_job_id"],
                    "revision": session["revision"],
                },
            )
            result, _, _ = self._evaluate(
                connection,
                session=session,
                job=job,
                catalog=catalog,
                definitions=definitions,
            )
            gate_id = self._insert_gate(connection, session=session, result=result)
            connection.execute(
                """
                UPDATE review_session SET status = ?, gate_status = ?, gate_result_hash = ?,
                    decision_set_hash = ?, updated_at = ? WHERE id = ?
                """,
                (
                    "READY_TO_CONFIRM" if result["status"] == "PASS" else "IN_REVIEW",
                    result["status"],
                    result["result_hash"],
                    result["decision_set_hash"],
                    utc_now(),
                    session_id,
                ),
            )
            self.database._record_event_in_connection(
                connection,
                event_type=(
                    "REVIEW_GATE_PASSED" if result["status"] == "PASS" else "REVIEW_GATE_BLOCKED"
                ),
                entity_type="review_session",
                entity_id=session_id,
                actor=safe_actor,
                detail={
                    "conversion_job_id": session["conversion_job_id"],
                    "review_gate_run_id": gate_id,
                    "result_hash": result["result_hash"],
                    "blocking_issue_count": result["blocking_issue_count"],
                    "decision_set_hash": result["decision_set_hash"],
                },
            )
        return {"gate_run_id": gate_id, **result}

    def confirm(
        self,
        session_id: str,
        *,
        snapshot_name: str,
        reviewer: str,
        reason: str,
        expected_revision: int,
        expected_candidate_set_hash: str,
        expected_decision_set_hash: str,
        run_after_confirm: bool,
        generate_charts: bool,
    ) -> dict[str, Any]:
        safe_name = str(snapshot_name).strip()[:160]
        safe_reviewer = str(reviewer).strip()[:120]
        safe_reason = str(reason).strip()[:1000]
        if not safe_name or not safe_reviewer or not safe_reason:
            raise ValueError("快照名称、确认人和确认理由均不能为空")
        try:
            with self.database.transaction() as connection:
                session, job, catalog, definitions = self._load_session_context(
                    connection, session_id
                )
                if session["status"] not in EDITABLE_SESSION_STATUSES:
                    if session["status"] == "CONFIRMED" and session.get("confirmed_snapshot_id"):
                        return {
                            "snapshot_id": session["confirmed_snapshot_id"],
                            "created": False,
                            "run_id": None,
                            "gate": None,
                        }
                    raise ValueError("复核会话当前不能确认")
                if int(session["revision"]) != int(expected_revision):
                    raise ReviewRevisionConflict(self.get_session(session_id))
                current_candidate_hash = self._candidate_set_hash(connection, job)
                decisions = self._effective_decisions(connection, session_id)
                current_decision_hash = self._decision_set_hash(decisions)
                if current_candidate_hash != str(expected_candidate_set_hash):
                    raise ValueError("候选集合哈希不匹配，请刷新后重新确认")
                if current_decision_hash != str(expected_decision_set_hash):
                    raise ValueError("复核决定集合哈希不匹配，请刷新后重新确认")
                self.database._record_event_in_connection(
                    connection,
                    event_type="REVIEW_CONFIRM_STARTED",
                    entity_type="review_session",
                    entity_id=session_id,
                    actor=safe_reviewer,
                    detail={
                        "conversion_job_id": session["conversion_job_id"],
                        "revision": session["revision"],
                        "candidate_set_hash": current_candidate_hash,
                        "decision_set_hash": current_decision_hash,
                    },
                )
                result, payload, decisions = self._evaluate(
                    connection,
                    session=session,
                    job=job,
                    catalog=catalog,
                    definitions=definitions,
                )
                gate_id = self._insert_gate(connection, session=session, result=result)
                if result["status"] != "PASS" or payload is None:
                    raise ValueError("最终质量门禁未通过，不能创建快照")
                payload_text = canonical_json(payload)
                payload_hash = json_sha256(payload)
                existing = connection.execute(
                    "SELECT id FROM input_snapshot WHERE payload_sha256 = ?", (payload_hash,)
                ).fetchone()
                created = existing is None
                if existing is None:
                    snapshot_id = self.database._insert_snapshot_in_connection(
                        connection,
                        payload,
                        name=safe_name,
                        source_path=f"review://{session_id}",
                        payload_text=payload_text,
                        payload_hash=payload_hash,
                        actor=safe_reviewer,
                    )
                else:
                    snapshot_id = str(existing["id"])
                confirmed_at = utc_now()
                connection.execute(
                    """
                    INSERT OR IGNORE INTO input_snapshot_provenance(
                        snapshot_id, conversion_job_id, converter_version,
                        mapping_profile_id, mapping_version, mapping_sha256,
                        contract_id, contract_version, contract_sha256,
                        source_manifest_json, case_sha256, confirmed_by, confirmed_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        snapshot_id,
                        job["id"],
                        job["converter_version"],
                        job["profile_id"],
                        job["profile_version"],
                        job["profile_sha256"],
                        job["contract_id"],
                        job["contract_version"],
                        job["contract_sha256"],
                        job["source_manifest_json"] or "{}",
                        payload_hash,
                        safe_reviewer,
                        confirmed_at,
                    ),
                )
                provenance_id = f"RPROV-{uuid4()}"
                decision_summary = [
                    {
                        "review_item_key": row["review_item_key"],
                        "field_id": row["field_id"],
                        "entity_id": row.get("entity_id"),
                        "action": row["action"],
                        "selected_candidate_id": row.get("selected_candidate_id"),
                        "decision_sha256": row["decision_sha256"],
                    }
                    for row in decisions
                ]
                connection.execute(
                    """
                    INSERT INTO input_snapshot_review_provenance(
                        id, snapshot_id, conversion_job_id, review_session_id,
                        review_gate_run_id, source_manifest_hash, candidate_set_hash,
                        decision_set_hash, mapping_version, mapping_sha256,
                        contract_version, contract_sha256, extraction_model_version,
                        ocr_model_version, prompt_version, rule_version,
                        decision_summary_json, confirmed_by, confirmation_reason, confirmed_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        provenance_id,
                        snapshot_id,
                        job["id"],
                        session_id,
                        gate_id,
                        session["source_manifest_hash"],
                        current_candidate_hash,
                        current_decision_hash,
                        job["profile_version"],
                        job["profile_sha256"],
                        job["contract_version"],
                        job["contract_sha256"],
                        job["extraction_model_version"],
                        job["ocr_model_version"],
                        _json(job["stage4_metrics_json"], {}).get("prompt_version"),
                        f"{ASSEMBLY_RULE_VERSION};{GATE_RULE_VERSION}",
                        canonical_json(decision_summary),
                        safe_reviewer,
                        safe_reason,
                        confirmed_at,
                    ),
                )
                connection.execute(
                    """
                    UPDATE review_session SET status = 'CONFIRMED', gate_status = 'PASS',
                        gate_result_hash = ?, decision_set_hash = ?,
                        confirmed_snapshot_id = ?, confirmed_at = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        result["result_hash"],
                        current_decision_hash,
                        snapshot_id,
                        confirmed_at,
                        confirmed_at,
                        session_id,
                    ),
                )
                connection.execute(
                    """
                    UPDATE conversion_job SET status = 'CONFIRMED', progress = 100,
                        status_message = '人工复核通过并写入不可变快照', snapshot_id = ?,
                        confirmed_by = ?, confirmed_at = ?, revision = revision + 1
                    WHERE id = ? AND status IN ('BLOCKED','READY_FOR_CONFIRMATION','CONFIRMED')
                    """,
                    (snapshot_id, safe_reviewer, confirmed_at, job["id"]),
                )
                run_id = None
                if run_after_confirm:
                    run_id = (
                        f"RUN-{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}-"
                        f"{uuid4().hex[:8]}"
                    )
                    connection.execute(
                        """
                        INSERT INTO calculation_run(
                            id, snapshot_id, status, input_sha256, created_at
                        )
                        VALUES (?, ?, 'QUEUED', ?, ?)
                        """,
                        (run_id, snapshot_id, payload_hash, utc_now()),
                    )
                    self.database._record_event_in_connection(
                        connection,
                        event_type="RUN_QUEUED",
                        entity_type="calculation_run",
                        entity_id=run_id,
                        actor=safe_reviewer,
                        detail={
                            "snapshot_id": snapshot_id,
                            "review_session_id": session_id,
                            "targets": session["target_node_ids"],
                            "generate_charts": bool(generate_charts),
                        },
                    )
                self.database._record_event_in_connection(
                    connection,
                    event_type="REVIEW_CONFIRMED",
                    entity_type="review_session",
                    entity_id=session_id,
                    actor=safe_reviewer,
                    detail={
                        "conversion_job_id": job["id"],
                        "snapshot_id": snapshot_id,
                        "snapshot_created": created,
                        "payload_sha256": payload_hash,
                        "candidate_set_hash": current_candidate_hash,
                        "decision_set_hash": current_decision_hash,
                        "review_gate_run_id": gate_id,
                        "provenance_id": provenance_id,
                        "run_id": run_id,
                    },
                )
            return {
                "snapshot_id": snapshot_id,
                "created": created,
                "run_id": run_id,
                "targets": session["target_node_ids"],
                "generate_charts": bool(generate_charts),
                "gate": result,
            }
        except ReviewRevisionConflict:
            raise
        except Exception as exc:
            self.database.record_event(
                event_type="REVIEW_CONFIRM_FAILED",
                entity_type="review_session",
                entity_id=session_id,
                actor=safe_reviewer or "local-reviewer",
                detail={"error_type": type(exc).__name__, "message": str(exc)[:300]},
            )
            raise

    def get_evidence(self, session_id: str, evidence_id: str) -> dict[str, Any]:
        with self.database.session() as connection:
            session, _, _, _ = self._load_session_context(connection, session_id)
            row = connection.execute(
                """
                SELECT link.evidence_json, link.candidate_id, candidate.payload_json
                FROM candidate_evidence_link link
                JOIN candidate_field candidate
                  ON candidate.job_id = link.job_id AND candidate.candidate_id = link.candidate_id
                WHERE link.job_id = ? AND link.evidence_id = ?
                ORDER BY link.candidate_id LIMIT 1
                """,
                (session["conversion_job_id"], evidence_id),
            ).fetchone()
            if row is None:
                raise KeyError(f"证据不属于当前复核任务：{evidence_id}")
            evidence = _json(row["evidence_json"], {})
            candidate = _json(row["payload_json"], {})
            location = evidence.get("location") or {}
            source_id = str(location.get("file_id") or "")
            source = connection.execute(
                """
                SELECT id, relative_path, original_file_name, detected_media_type,
                       byte_count, sha256, parser_id, parser_version,
                       parse_quality_json, parsed_at
                FROM conversion_source WHERE job_id = ? AND id = ?
                """,
                (session["conversion_job_id"], source_id),
            ).fetchone()
            if source is None:
                raise KeyError("证据登记的源文件不存在或不属于当前任务")
        source_doc = dict(source)
        source_doc["parse_quality"] = _json(source_doc.pop("parse_quality_json"), {})
        media_type = str(source_doc.get("detected_media_type") or "application/octet-stream")
        versions = candidate.get("model_or_rule_versions") or {}
        return {
            "evidence_id": evidence_id,
            "candidate_id": row["candidate_id"],
            "conversion_job_id": session["conversion_job_id"],
            "source_id": source_id,
            "sanitized_file_name": source_doc.get("relative_path")
            or source_doc.get("original_file_name"),
            "media_type": media_type,
            "sha256": source_doc.get("sha256"),
            "parser_id": source_doc.get("parser_id"),
            "parser_version": source_doc.get("parser_version"),
            "extraction_method": candidate.get("extraction_method"),
            "provider_id": versions.get("provider_id") or versions.get("provider"),
            "model_version": versions.get("model_version") or versions.get("model"),
            "prompt_version": versions.get("prompt_version") or versions.get("prompt"),
            "model_or_rule_versions": versions,
            "page_number": location.get("page_number") or location.get("page"),
            "sheet_name": location.get("sheet_name"),
            "cell_range": location.get("cell_range")
            or (
                f"R{location.get('row')}C{location.get('column')}"
                if location.get("row") is not None
                else None
            ),
            "bounding_box": location.get("bounding_box") or location.get("bbox"),
            "location": location,
            "original_text": evidence.get("excerpt") or location.get("cell_text"),
            "normalized_text": candidate.get("normalized_value"),
            "confidence": candidate.get("confidence"),
            "parsed_at": source_doc.get("parsed_at"),
            "source_quality": source_doc.get("parse_quality"),
            "preview_url": (
                f"/admin/api/review-sessions/{session_id}/evidence/{evidence_id}/preview"
            ),
        }

    def evidence_preview(
        self, session_id: str, evidence_id: str
    ) -> tuple[str, bytes, dict[str, str]]:
        metadata = self.get_evidence(session_id, evidence_id)
        source = self.database.conversion_source_content(
            str(metadata["source_id"]),
            job_id=str(metadata["conversion_job_id"]),
            allowed_statuses={"PARSED", "PARSE_FAILED", "READY_FOR_PARSE"},
        )
        content = bytes(source["content"])
        media_type = str(metadata["media_type"])
        headers = {
            "Content-Disposition": "inline",
            "Content-Security-Policy": (
                "default-src 'none'; img-src 'self' data:; style-src 'unsafe-inline'"
            ),
        }
        location = metadata.get("location") or {}
        bbox = location.get("bounding_box") or location.get("bbox")
        if media_type.startswith("image/"):
            try:
                image = Image.open(io.BytesIO(content))
                highlighted = _highlight_bbox(image, bbox, location.get("image_size"))
                headers["X-QRA-Preview-Source"] = media_type
                return "image/png", _png_bytes(highlighted), headers
            except Exception:
                return media_type, content, headers
        if media_type == "application/pdf":
            try:
                page_number = max(1, int(metadata.get("page_number") or 1))
                with pdfplumber.open(io.BytesIO(content)) as document:
                    if page_number > len(document.pages):
                        raise ValueError("证据页码超出PDF范围")
                    page = document.pages[page_number - 1]
                    image = page.to_image(resolution=144, antialias=True).original
                source_size = location.get("page_size") or [page.width, page.height]
                highlighted = _highlight_bbox(image, bbox, source_size)
                headers["X-QRA-Preview-Source"] = "application/pdf"
                headers["X-QRA-Preview-Page"] = str(page_number)
                return "image/png", _png_bytes(highlighted), headers
            except Exception:
                return media_type, content, headers
        context: dict[str, Any] = {"metadata": metadata, "rows": []}
        try:
            row_number = max(1, int(location.get("row") or 1))
            if media_type in {"text/csv", "application/csv"} or str(
                metadata["sanitized_file_name"]
            ).lower().endswith(".csv"):
                text = content.decode("utf-8-sig", errors="replace")
                rows = list(csv.reader(io.StringIO(text)))
                start, end = max(0, row_number - 3), min(len(rows), row_number + 2)
                context["rows"] = [
                    {"row": index + 1, "values": rows[index]} for index in range(start, end)
                ]
            elif str(metadata["sanitized_file_name"]).lower().endswith(".xlsx"):
                workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                sheet_name = str(location.get("sheet_name") or workbook.sheetnames[0])
                sheet = workbook[sheet_name]
                start, end = max(1, row_number - 2), min(sheet.max_row, row_number + 2)
                context["rows"] = [
                    {
                        "row": index,
                        "values": [cell.value for cell in sheet[index]],
                    }
                    for index in range(start, end + 1)
                ]
            elif str(metadata["sanitized_file_name"]).lower().endswith(".xls"):
                workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
                sheet_name = str(location.get("sheet_name") or workbook.sheet_names()[0])
                sheet = workbook.sheet_by_name(sheet_name)
                start, end = max(0, row_number - 3), min(sheet.nrows, row_number + 2)
                context["rows"] = [
                    {"row": index + 1, "values": sheet.row_values(index)}
                    for index in range(start, end)
                ]
        except Exception as exc:
            context["context_error"] = f"{type(exc).__name__}: {exc}"
        return (
            "application/json; charset=utf-8",
            (json.dumps(context, ensure_ascii=False, indent=2) + "\n").encode("utf-8"),
            headers,
        )


__all__ = [
    "DECISION_ACTIONS",
    "EDITABLE_SESSION_STATUSES",
    "REVIEW_SERVICE_VERSION",
    "ReviewRevisionConflict",
    "ReviewService",
]
